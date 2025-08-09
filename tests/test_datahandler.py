import sys
from pathlib import Path
import types

import openpyxl

sys.path.append(str(Path(__file__).resolve().parents[1]))

# Stub external dependencies used during import
sys.modules.setdefault("openai", types.SimpleNamespace())

stub_got = types.ModuleType("GOT")
stub_got.get_output = lambda: ["a", "b", "c", "d", "e", "f"]

stub_elementen = types.ModuleType("Elementen")
for name in [
    "count_onderkasten_kolomkasten",
    "count_Hangkasten",
    "count_plinten",
    "count_Kroonlijsten",
    "count_Lichtlijsten",
    "count_Werkbladen",
    "count_Spoelbak_kraan",
    "count_Kookplaat",
    "count_Dampkap",
    "count_Oven",
    "count_Microgolfoven",
    "count_koelkast",
    "count_Vaatwasser",
    "count_Passtukken",
    "count_Deur_vaatwasser",
    "count_Achterwand",
    "count_Zij_steunwand",
    "count_Andere",
    "count_Verlichting",
]:
    setattr(stub_elementen, name, lambda *_, **__: 1)

# Attributes referenced by Datahandler
stub_elementen.filename = ""
stub_elementen.target_words_onderkasten_kolomkasten = []
stub_elementen.target_words_Hangkasten = []
stub_elementen.target_words_plinten = []
stub_elementen.target_words_Kroonlijsten = []
stub_elementen.target_words_Lichtlijsten = []
stub_elementen.target_words_Werkbladen = []
stub_elementen.target_words_Spoelbak_kraan = []
stub_elementen.target_words_Kookplaat = []
stub_elementen.target_words_Dampkap = []
stub_elementen.target_words_Oven = []
stub_elementen.target_words_Microgolfoven = []
stub_elementen.target_words_koelkast = []
stub_elementen.target_words_Vaatwasser = []
stub_elementen.target_words_Passtukken = []
stub_elementen.target_words_Deur_vaatwasser = []
stub_elementen.target_words_Achterwand = []
stub_elementen.target_words_Zij_steunwand = []
stub_elementen.target_words_Andere = []
stub_elementen.target_words_Verlichting = []

sys.modules["GOT"] = stub_got
sys.modules["Elementen"] = stub_elementen

import Datahandler  # noqa: E402


def test_generate_opmetingsaanvraag(tmp_path):
    template = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    wb.save(template)

    Datahandler.generate_opmetingsaanvraag(str(template))
    wb2 = openpyxl.load_workbook(template)
    ws = wb2.active
    assert ws["C6"].value == "a"
    assert ws["E15"].value == 1
