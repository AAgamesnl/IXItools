import openpyxl
import re
import math
filename = "Data.xlsx"

def count_onderkasten_kolomkasten(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    if "wand-hoekkast" not in cell_value.lower() and "kleinere dieptes bij onderkast" not in cell_value.lower():
                        for word in words:
                            if word.lower() in cell_value.lower():
                                word_counts[word] += 1

    # Calculate the total count
    total_count_onderkasten_kolomkasten = sum(word_counts.values())

    # Return the total count
    return total_count_onderkasten_kolomkasten

target_words_onderkasten_kolomkasten = ["Onderkast", "Hoekkast", "Ladekast", "Kookplaatkast", "Spoelkast", "Ombouwkast", "Servies- / Voorraadkast"]
total_count_onderkasten_kolomkasten = count_onderkasten_kolomkasten(filename, target_words_onderkasten_kolomkasten)

def count_Hangkasten(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Hangkasten = sum(word_counts.values())

    # Return the total count
    return total_count_Hangkasten
target_words_Hangkasten = ["Wandkast", "Wand-hoekkast"]
total_count_Hangkasten = count_Hangkasten(filename, target_words_Hangkasten)

def count_plinten(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for 'plinten'
    plinten_count = 0

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            plinten_count += 1

    # Calculate the total count
    total_count_plinten = plinten_count / 2

    # Round down to the nearest whole number
    total_count_plinten = math.floor(total_count_plinten)

    # If there were any 'plinten', make sure the total is at least 1
    if plinten_count > 0 and total_count_plinten < 1:
        total_count_plinten = 1

    # Return the total count
    return total_count_plinten

target_words_plinten = ["SB15", "SB7", "SB20"]
total_count_plinten = count_plinten(filename, target_words_plinten)

def count_Kroonlijsten(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Kroonlijsten = sum(word_counts.values())

    # Return the total count
    return total_count_Kroonlijsten
target_words_Kroonlijsten = ["Kroonlijst"]
total_count_Kroonlijsten = count_Kroonlijsten(filename, target_words_Kroonlijsten)

def count_Lichtlijsten(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Kroonlijsten = sum(word_counts.values())

    # Return the total count
    return total_count_Kroonlijsten
target_words_Lichtlijsten = ["Lichtlijsten"]
total_count_Lichtlijsten = count_Lichtlijsten(filename, target_words_Lichtlijsten)

def count_Werkbladen(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Werkbladen = sum(word_counts.values())

    # Return the total count
    return total_count_Werkbladen
target_words_Werkbladen = ["Werkblad APD", "Werkblad APN"]
total_count_Werkbladen = count_Werkbladen(filename, target_words_Werkbladen)

def count_Spoelbak_kraan(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Spoelbak_kraan = sum(word_counts.values())
    if total_count_Spoelbak_kraan >= 1:
        total_count_Spoelbak_kraan = 3
    # Return the total count
    return total_count_Spoelbak_kraan
target_words_Spoelbak_kraan = ["Spoelkast", "kraan", "SPOELBAK"]
total_count_Spoelbak_kraan = count_Spoelbak_kraan(filename, target_words_Spoelbak_kraan)

def count_Kookplaat(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Kookplaat = sum(word_counts.values())*2
    # Return the total count
    return total_count_Kookplaat
target_words_Kookplaat = ["Kookplaatkast", "kookplaat"]
total_count_Kookplaat = count_Kookplaat(filename, target_words_Kookplaat)

def count_Dampkap(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Dampkap = sum(word_counts.values())
    if total_count_Dampkap >=1:
        total_count_Dampkap = 1
    # Return the total count
    return total_count_Dampkap
target_words_Dampkap = ["vlakscherm", "Wandschouwkap", "Wanddampkap", "dampkap"]
total_count_Dampkap = count_Dampkap(filename, target_words_Dampkap)

def count_Oven(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Oven = sum(word_counts.values())
    if total_count_Oven >=1:
        total_count_Oven = 1
    # Return the total count
    return total_count_Oven
target_words_Oven = ["oven"]
total_count_Oven = count_Oven(filename, target_words_Oven)

def count_Microgolfoven(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Microgolfoven = sum(word_counts.values())
    if total_count_Microgolfoven >=1:
        total_count_Microgolfoven = 1
    # Return the total count
    return total_count_Microgolfoven
target_words_Microgolfoven = ["magnetron"]
total_count_Microgolfoven = count_Microgolfoven(filename, target_words_Microgolfoven)

def count_koelkast(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_koelkast = sum(word_counts.values())*2
    # Return the total count
    return total_count_koelkast
target_words_koelkast = ["Ombouwkast koel / vries", "Ombouwkast koelautomaat", "GD1"]
total_count_koelkast = count_koelkast(filename, target_words_koelkast)

def count_Vaatwasser(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Vaatwasser = sum(word_counts.values())
    # Return the total count
    return total_count_Vaatwasser
target_words_Vaatwasser = ["Deurfront voor", "Deurfront in", "Doorlopend deurfront"]
total_count_Vaatwasser = count_Vaatwasser(filename, target_words_Vaatwasser)

def count_Passtukken(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Passtukken = sum(word_counts.values())
    # Return the total count
    return total_count_Passtukken
target_words_Passtukken = ["passtuk"]
total_count_Passtukken = count_Passtukken(filename, target_words_Passtukken)

def count_afdekbodems(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_afdekbodems = sum(word_counts.values())
    # Return the total count
    return total_count_afdekbodems
target_words_afdekbodems = ["afdekbodem"]
total_count_afdekbodems = count_afdekbodems(filename, target_words_afdekbodems)

def count_Deur_vaatwasser(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            if word.lower() == "deurfront in":
                                word_counts[word] += 2  # Increment count by 2 for "Deurfront in"
                            else:
                                word_counts[word] += 1  # Increment count by 1 for other words

    # Calculate the total count
    total_count_Deur_vaatwasser = sum(word_counts.values())

    # Return the total count
    return total_count_Deur_vaatwasser
target_words_Deur_vaatwasser = ["Deurfront voor", "Deurfront in", "Doorlopend deurfront"]
total_count_Deur_vaatwasser = count_Deur_vaatwasser(filename, target_words_Deur_vaatwasser)

def count_Achterwand(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        pattern = r"\b{}\b".format(re.escape(word))
                        if re.search(pattern, cell_value, re.IGNORECASE):
                            if word.lower() == "achterwandbekledingen":
                                continue  # Skip counting "achterwandbekledingen"
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Achterwand = sum(word_counts.values())

    # Return the total count
    return total_count_Achterwand
target_words_Achterwand = ["Achterwandbekleding"]
total_count_Achterwand = count_Achterwand(filename, target_words_Achterwand)


def count_Zij_steunwand(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Zij_steunwand = sum(word_counts.values())
    # Return the total count
    return total_count_Zij_steunwand
target_words_Zij_steunwand = ["HWK10", "HWK16", "HWK25", "HWK50", "WA10", "WA16", "WA25", "WA50", "WW10", "WW16", "WW25", "WW50", "WR10", "WR16", "WR25", "WR50", "Steunvoet"]
total_count_Zij_steunwand = count_Zij_steunwand(filename, target_words_Zij_steunwand)

def count_Andere(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Andere = sum(word_counts.values())
    if total_count_Andere >= 1:
        total_count_Andere = 5
    # Return the total count
    return total_count_Andere
target_words_Andere = ["LineN"]
total_count_Andere = count_Andere(filename, target_words_Andere)

def count_Verlichting(filename, words):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Initialize a counter for each word
    word_counts = {word: 0 for word in words}

    # Iterate over all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Iterate over all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value contains any of the target words
                    cell_value = str(cell.value)
                    for word in words:
                        if word.lower() in cell_value.lower():
                            word_counts[word] += 1

    # Calculate the total count
    total_count_Verlichting = sum(word_counts.values())
    # Return the total count
    return total_count_Verlichting
target_words_Verlichting = ["LNDERB"]
total_count_Verlichting = count_Verlichting(filename, target_words_Verlichting)