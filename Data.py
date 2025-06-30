import PyPDF2
import re
from openpyxl import Workbook, load_workbook
from tkinter import filedialog

# Ask the user to select a PDF file
pdf_file_path = filedialog.askopenfilename(title="Selecteer de bestelbon",
                                           filetypes=[("PDF files", "*.pdf")])

# Lees het PDF-bestand en extraheer de tekst
with open(pdf_file_path, 'rb') as pdf_file:
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    text = ''.join(page.extract_text() for page in pdf_reader.pages)

# Voeg spatie toe tussen woorden op verschillende regels
new_lines = [text.split('\n')[0]]
for line in text.split('\n')[1:]:
    if line and line[0].islower():
        new_lines[-1] += ' ' + line
    else:
        new_lines.append(line)

# Filter de regels
filtered_lines = [line for line in new_lines if 'Netto afmetingen :' not in line and 'Line n' not in line]

# Zoek de index van de regel met de totale prijs van de keuken
total_price_index = next(i for i, line in enumerate(filtered_lines) if 'Totale prijs van uw ingerichte keuken incl. BTW:' in line)

# Schrijf de regels naar een Excel-bestand
workbook = Workbook()
worksheet = workbook.active
for i, line in enumerate(filtered_lines[:total_price_index + 1], 1):
    worksheet.cell(row=i, column=1, value=line)

workbook.save('Data.xlsx')