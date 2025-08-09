"""Utilities for populating the 'Opmetingsaanvraag' Excel sheet.

This module generates the element list by combining output from
:mod:`GOT` and :mod:`Elementen`. The heavy GUI interactions are isolated in
:func:`main` so that the core generation logic remains testable and free of
side effects on import.
"""

from __future__ import annotations

import os
import shutil
from pathlib import Path
from typing import Sequence

import openpyxl
import GOT
import Elementen
import tkinter as tk
from tkinter import filedialog, messagebox

# Mapping for dynamic cell placement of header information
CELL_MAPPING: dict[int, str] = {
    1: "C6",
    2: "C7",
    3: "C9",
    4: "C10",
    5: "D12",
    6: "G6",
}


def _write_header(
    ws: openpyxl.worksheet.worksheet.Worksheet, data: Sequence[str]
) -> None:
    """Write header data to predefined cells."""
    for i, line in enumerate(data, 1):
        if i in CELL_MAPPING:
            ws[CELL_MAPPING[i]] = line


def _collect_counts() -> dict[str, int]:
    """Collect element counts from :mod:`Elementen`."""
    return {
        "E15": Elementen.count_onderkasten_kolomkasten(
            Elementen.filename, Elementen.target_words_onderkasten_kolomkasten
        ),
        "E16": Elementen.count_Hangkasten(
            Elementen.filename, Elementen.target_words_Hangkasten
        ),
        "E17": Elementen.count_plinten(
            Elementen.filename, Elementen.target_words_plinten
        ),
        "E18": Elementen.count_Kroonlijsten(
            Elementen.filename, Elementen.target_words_Kroonlijsten
        ),
        "E19": Elementen.count_Lichtlijsten(
            Elementen.filename, Elementen.target_words_Lichtlijsten
        ),
        "E20": Elementen.count_Werkbladen(
            Elementen.filename, Elementen.target_words_Werkbladen
        ),
        "E21": Elementen.count_Spoelbak_kraan(
            Elementen.filename, Elementen.target_words_Spoelbak_kraan
        ),
        "E22": Elementen.count_Kookplaat(
            Elementen.filename, Elementen.target_words_Kookplaat
        ),
        "E23": Elementen.count_Dampkap(
            Elementen.filename, Elementen.target_words_Dampkap
        ),
        "E24": Elementen.count_Oven(Elementen.filename, Elementen.target_words_Oven),
        "E25": Elementen.count_Microgolfoven(
            Elementen.filename, Elementen.target_words_Microgolfoven
        ),
        "E26": Elementen.count_koelkast(
            Elementen.filename, Elementen.target_words_koelkast
        ),
        "E27": Elementen.count_Vaatwasser(
            Elementen.filename, Elementen.target_words_Vaatwasser
        ),
        "E30": Elementen.count_Passtukken(
            Elementen.filename, Elementen.target_words_Passtukken
        ),
        "E33": Elementen.count_Deur_vaatwasser(
            Elementen.filename, Elementen.target_words_Deur_vaatwasser
        ),
        "E34": Elementen.count_Achterwand(
            Elementen.filename, Elementen.target_words_Achterwand
        ),
        "E35": Elementen.count_Zij_steunwand(
            Elementen.filename, Elementen.target_words_Zij_steunwand
        ),
        "E37": Elementen.count_Andere(
            Elementen.filename, Elementen.target_words_Andere
        ),
        "E38": Elementen.count_Verlichting(
            Elementen.filename, Elementen.target_words_Verlichting
        ),
    }


def generate_opmetingsaanvraag(template_path: str = "Opmetingsaanvraag.xlsx") -> Path:
    """Populate ``template_path`` with project data and return the saved path."""
    template = Path(template_path)
    wb = openpyxl.load_workbook(template)
    ws = wb.active

    _write_header(ws, GOT.get_output())

    for cell, value in _collect_counts().items():
        ws[cell] = value

    wb.save(template)
    return template


def main() -> None:
    """Generate the spreadsheet and optionally open or save it elsewhere."""
    output_path = generate_opmetingsaanvraag()

    root = tk.Tk()
    root.withdraw()  # hide the empty tkinter window

    action = messagebox.askyesno(
        "Elementenlijst klaar",
        "Wil je het bestand openen? Klik op 'Nee' om de elementenlijst op te slaan",
    )
    if action:
        os.system(f"start excel.exe {output_path}")
    else:
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=str(output_path.name),
            filetypes=[("Excel files", "*.xlsx")],
        )
        if file_path:
            try:
                shutil.move(str(output_path), file_path)
                print(f"Elementen lijst opgeslagen in {file_path}")
            except Exception as exc:  # pragma: no cover - interactive path
                print(
                    "Er is een fout met het opslaan het de elementenlijst. "
                    f"Contacteer Ayoub: {exc}"
                )
        else:
            messagebox.showinfo("Das pech, lijst weg.", "Geen locatie aangegeven")


if __name__ == "__main__":  # pragma: no cover - manual execution
    main()
